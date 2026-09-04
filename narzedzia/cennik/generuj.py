#!/usr/bin/env python3
"""Składa app/src/data/price-data.js z raportu magazynowego.

    python3 narzedzia/cennik/generuj.py raport.xlsx [--sprawdz]

Raport daje ceny netto i ilości. Reszta — który silnik pasuje do którego
wariantu, lista falowników, osprzętu i masy do przesyłki — pochodzi
z narzedzia/cennik/katalog.json.

Lista wariantów pochodzi z tabeli doborowej aplikacji (catalog-data.js),
a nie z magazynu: co aplikacja potrafi zaproponować, musi mieć wiersz
w cenniku, inaczej znika z wyników bez żadnego śladu.

Stany zapisujemy jako 0/1 — liczba sztuk celowo nie opuszcza magazynu.
Z --sprawdz plik nie jest zapisywany, wypisywane jest samo podsumowanie.
"""
import json, re, sys, unicodedata
from datetime import datetime, date
from pathlib import Path

KORZEN = Path(__file__).resolve().parents[2]
CENNIK = KORZEN / 'app/src/data/price-data.js'
# poniżej tylu wariantów albo tylu pozycji w raporcie uznajemy plik za uszkodzony
MIN_WARIANTOW, MIN_POZYCJI = 1600, 400


def norm(s):
    """Klucz porównania kodów: bez ogonków, wielkimi literami, pojedyncze spacje.

    W raporcie ten sam towar bywa raz z „Ł", raz z „L", i z podwójną spacją."""
    s = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().upper()


def wczytaj_raport(sciezka):
    import openpyxl
    ws = openpyxl.load_workbook(sciezka, data_only=True)[
        openpyxl.load_workbook(sciezka, data_only=True).sheetnames[0]]
    naglowek = ws.cell(1, 1).value
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(naglowek))
    if not m:
        sys.exit(f'Nie znalazłem daty raportu w komórce A1: {naglowek!r}')
    dzien = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    poz = {}
    for r in range(8, ws.max_row + 1):
        kod = ws.cell(r, 1).value
        if not isinstance(kod, str) or not kod.strip():
            continue
        poz[norm(kod)] = {'ilosc': ws.cell(r, 3).value or 0,
                          'cena': ws.cell(r, 4).value,
                          'nazwa': str(ws.cell(r, 2).value or '')}
    return dzien, poz


def nazwa_handlowa(surowa):
    """Z opisu magazynowego zostawiamy to, co ma sens dla klienta."""
    s = re.sub(r'^Silnik( elektryczny)?\s+', '', surowa, flags=re.I)
    s = re.split(r'\s+IP55', s, flags=re.I)[0]
    return re.sub(r'\s*\([^()]*\)\s*$', '', s).strip()


def js(v):
    if v is None:
        return 'null'
    if isinstance(v, str):
        return "'" + v.replace('\\', '\\\\').replace("'", "\\'") + "'"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return repr(round(v, 4) if isinstance(v, float) else v)


# B34 to łapy + kołnierz B14, B35 to łapy + kołnierz B5 — pasują tam, gdzie tabela
# chce B14 albo B5; łapy są naddatkiem. Magazyn trzyma część silników pod tymi kodami.
ROWNOWAZNE_MOCOWANIA = {'B34': 'B14', 'B35': 'B5', 'B14': 'B14', 'B5': 'B5'}
BIEGUNY = {'2': 2800, '4': 1400, '6': 900}


def silniki_z_raportu(poz):
    """Indeks trójfazowych silników DKM z raportu: (kW|obroty|kołnierz) -> kody."""
    idx = {}
    for kod in poz:
        if '1 FAZOWY' in kod or 'DKM' not in kod:
            continue
        m = re.match(r'^(\d+(?:,\d+)?) ([246]) (\d{2,3})(B\d+)\b', kod)
        if not m or m.group(4) not in ROWNOWAZNE_MOCOWANIA:
            continue
        kw = float(m.group(1).replace(',', '.'))
        klucz = f'{kw:g}|{BIEGUNY[m.group(2)]}|{m.group(3)}{ROWNOWAZNE_MOCOWANIA[m.group(4)]}'
        idx.setdefault(klucz, []).append(kod)
    return idx


def kod_przekladni(box, iec, i):
    """Kody w magazynie: „DKM025 56B14 I20", przełożenie ułamkowe z przecinkiem."""
    prz = 'I' + (str(i).replace('.', ',') if '.' in str(i) else str(i))
    warianty = [f'{box} {iec} {prz}']
    # jedna przekładnia pasuje do ramki 100 i 112 — magazyn trzyma ją jako jeden kod
    m = re.match(r'^(100|112)(B\d+)$', iec)
    if m:
        warianty.append(f'{box} 100/112{m.group(2)} {prz}')
    return warianty


def main():
    arg = [a for a in sys.argv[1:] if not a.startswith('--')]
    if not arg:
        sys.exit(__doc__)
    tylko_sprawdz = '--sprawdz' in sys.argv
    dzien, poz = wczytaj_raport(arg[0])
    K = json.loads((KORZEN / 'narzedzia/cennik/katalog.json').read_text('utf-8'))
    warianty = json.loads((KORZEN / 'narzedzia/cennik/warianty.json').read_text('utf-8'))

    # data trafia na stronę jako „stan na…", więc nie wolno jej cofnąć starszym
    # raportem — inaczej klient zobaczyłby nieprawdziwą datę bez żadnego błędu
    if CENNIK.exists():
        m = re.search(r"updated:\s*'stan na (\d{2})\.(\d{2})\.(\d{4})'", CENNIK.read_text('utf-8'))
        if m:
            obecna = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            if dzien < obecna:
                print(f'Raport z {dzien:%d.%m.%Y} jest starszy niż cennik '
                      f'({obecna:%d.%m.%Y}) — zostawiam bez zmian.')
                sys.exit(0)

    if len(poz) < MIN_POZYCJI:
        sys.exit(f'Raport ma tylko {len(poz)} pozycji (próg {MIN_POZYCJI}) — nie ruszam cennika.')

    def z_raportu(kody):
        for k in kody if isinstance(kody, list) else [kody]:
            w = poz.get(norm(k))
            if w:
                return w
        return None

    surowe, uzyte_sku, bez_silnika, bez_przekladni = [], set(), [], []
    nieprzypisane = set()
    for box, iec, i, kw, rpm, silnik_kat in warianty:
        p = z_raportu(kod_przekladni(box, iec, i))
        # raport odświeża cenę tam, gdzie zna kod; poza tym obowiązuje lista cenowa
        cena_p = p['cena'] if p and p['cena'] is not None else K['przekladnie'].get(f'{box}|{iec}|{i}')
        stan_p = 1 if p and p['ilosc'] else 0
        if cena_p is None:
            bez_przekladni.append(f'{box} {iec} i{i}')

        sku = K['silniki'].get(f'{kw}|{rpm}|{iec}', '')
        s = z_raportu(sku) if sku else None
        cena_s = (s['cena'] if s and s['cena'] is not None
                  else K['silnikiCeny'].get(sku) if sku else None)
        stan_s = 1 if s and s['ilosc'] else 0
        if not sku:
            bez_silnika.append(f'{box} {iec} i{i} · {kw}kW {rpm}obr (katalog: silnik {silnik_kat})')
            nieprzypisane.add(f'{float(kw):g}|{rpm}|{iec}')
        else:
            uzyte_sku.add(sku)

        surowe.append((box, f'{box}|{iec}|{i}|{kw}|{rpm}', cena_p, stan_p, cena_s, stan_s, sku))

    # Ceny przekładni NIE uzupełniamy ceną korpusu. Wyglądało to kusząco — cena jest
    # w obrębie korpusu stała — ale tabela doborowa producenta wymienia kombinacje
    # kołnierza i przełożenia, których DKM nie ma w ofercie (np. DKM063 71B14 i20).
    # Uzupełnienie zamieniało im status z „zapytaj o cenę" na „dostawa 1–3 dni",
    # czyli z prośby o kontakt na obietnicę terminu dla czegoś, czego nie da się
    # zamówić. Cenę ma tylko to, co widzieliśmy jako realny produkt: pozycja
    # z raportu magazynowego albo wpis z listy cenowej w katalog.json.
    var = {}
    for box, klucz, cena_p, stan_p, cena_s, stan_s, sku in surowe:
        brak = cena_p is None or cena_s is None
        cena_z = None if brak else round(cena_p + cena_s, 2)
        stan_z = 0 if brak else (1 if stan_p and stan_s else 0)
        status = 2 if brak else (0 if stan_z else 1)
        var[klucz] = [cena_p, stan_p, cena_s, stan_s, cena_z, stan_z, status, sku]

    if len(var) < MIN_WARIANTOW:
        sys.exit(f'Wyszło tylko {len(var)} wariantów (próg {MIN_WARIANTOW}) — nie ruszam cennika.')

    nam = {}
    for sku in sorted(uzyte_sku):
        w = z_raportu(sku)
        nam[sku] = nazwa_handlowa(w['nazwa']) if w else K['nazwyZapas'].get(sku, sku)

    opt = {}
    for klucz, (kod, cena_zapas) in sorted(K['osprzet'].items()):
        w = z_raportu(kod)
        opt[klucz] = [w['cena'] if w else cena_zapas, 1 if w and w['ilosc'] else 0]

    inv = []
    for nazwa, seria, moc, fazy, cena_zapas in K['falowniki']:
        w = z_raportu(nazwa)
        stan = 1 if w and w['ilosc'] else 0
        inv.append([nazwa, seria, moc, fazy, w['cena'] if w else cena_zapas, stan, 0 if stan else 1])

    m1f = {}
    for klucz, (kod, nazwa, cena_zapas) in sorted(K['m1f'].items()):
        w = z_raportu(kod)
        m1f[klucz] = [kod, nazwa, w['cena'] if w else cena_zapas, 1 if w and w['ilosc'] else 0]

    na_stanie = sum(1 for w in var.values() if w[5])
    print(f'raport {dzien:%d.%m.%Y} · pozycji {len(poz)} · wariantów {len(var)} '
          f'· zestawów na stanie {na_stanie} · silników {len(nam)}')
    brak_ceny = sum(1 for w in var.values() if w[0] is None)
    if brak_ceny:
        print(f'  wariantów wciąż bez ceny przekładni: {brak_ceny}')
    if bez_silnika:
        print(f'  wariantów bez przypisanego silnika (status „zapytaj o cenę"): {len(bez_silnika)}')
        for x in sorted(set(bez_silnika))[:12]:
            print('    ', x)

    # nowy silnik w magazynie sam się nie przypisze — przypisania nie zgadujemy, bo pod
    # tą samą mocą i kołnierzem bywa kilka wykonań (np. zwykłe i HPS) o różnych cenach.
    # Zamiast tego wypisujemy kandydatów, żeby nie zniknęli po cichu.
    idx = silniki_z_raportu(poz)
    kandydaci = [(k, idx[k]) for k in sorted(nieprzypisane) if k in idx]
    if kandydaci:
        print(f'  DO SPRAWDZENIA — magazyn ma silnik DKM, którego nie ma w katalog.json: '
              f'{len(kandydaci)}')
        for klucz, kody in kandydaci:
            kw, rpm, iec = klucz.split('|')
            print(f'    {kw} kW · {rpm} obr · {iec} → ' + ', '.join(kody))

    if tylko_sprawdz:
        return

    # bez znacznika czasu generowania: przy tych samych danych plik ma wychodzić
    # bit w bit taki sam, inaczej automat commituje i publikuje codziennie, choć
    # nic się nie zmieniło. Kiedy plik powstał, mówi historia repozytorium.
    L = [f"""/* DKM — ceny netto i dostępność.
   Wygenerowane automatycznie z raportu magazynowego: {dzien:%d.%m.%Y}
   NIE EDYTOWAĆ RĘCZNIE — plik jest nadpisywany codziennie.

   Stany podawane są jako 0/1 (nie ma / jest) — liczba sztuk
   celowo nie opuszcza magazynu.
*/
window.DKM_PRICE = {{
  updated: 'stan na {dzien:%d.%m.%Y}',
  var: {{"""]
    L.append(',\n'.join(f"  {js(k)}: [{','.join(js(x) for x in w)}]" for k, w in var.items()))
    L.append('  },\n  nam: {')
    L.append(',\n'.join(f'  {js(k)}: {js(v)}' for k, v in nam.items()))
    L.append('  },\n  opt: {')
    L.append(',\n'.join(f"  {js(k)}: [{js(v[0])},{js(v[1])}]" for k, v in opt.items()))
    L.append('  },\n  inv: [')
    L.append(',\n'.join('  [' + ','.join(js(x) for x in w) + ']' for w in inv))
    L.append('  ],\n  m1f: {')
    L.append(',\n'.join(f"  {js(k)}: [" + ','.join(js(x) for x in v) + ']' for k, v in m1f.items()))
    L.append('  },\n  wt: {')
    grupy = []
    for nazwa_g, tresc in K['wt'].items():
        w = ',\n'.join(f'      {js(k)}: {js(v)}' for k, v in tresc.items())
        grupy.append(f'    {nazwa_g}: {{\n{w}\n    }}')
    L.append(',\n'.join(grupy))
    L.append('  }\n};\n')
    CENNIK.write_text('\n'.join(L), 'utf-8')
    print(f'zapisano {CENNIK.relative_to(KORZEN)}')


if __name__ == '__main__':
    main()
